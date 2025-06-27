import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import sys
import asyncio
import re

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from agents.locate_agent import create_location_agent

def parse_distance_info(response_text):
    """解析距离信息，提取关键数据"""
    # 尝试从响应中提取距离和时间信息
    distance_match = re.search(r'(\d+\.?\d*)公里', response_text)
    time_match = re.search(r'(\d+)分钟', response_text)
    
    if distance_match and time_match:
        return {
            'distance': f"{distance_match.group(1)}公里",
            'duration': f"{time_match.group(1)}分钟",
            'distance_km': float(distance_match.group(1)),
            'duration_min': int(time_match.group(1)),
            'success': True
        }
    else:
        return {
            'distance': '未知',
            'duration': '未知',
            'distance_km': None,
            'duration_min': None,
            'success': False,
            'raw_response': response_text
        }

async def calculate_warehouse_to_warehouse_distance(agent, warehouse1, warehouse2, max_retries=3):
    """计算两个仓库之间的距离"""
    # 使用经纬度坐标进行计算
    warehouse1_lng = warehouse1['location']['longitude']
    warehouse1_lat = warehouse1['location']['latitude']
    warehouse1_location = f"{warehouse1_lng},{warehouse1_lat}"
    
    warehouse2_lng = warehouse2['location']['longitude']
    warehouse2_lat = warehouse2['location']['latitude']
    warehouse2_location = f"{warehouse2_lng},{warehouse2_lat}"
    
    query = f"从{warehouse1_location}到{warehouse2_location}的车辆行驶距离"
    
    for attempt in range(max_retries):
        try:
            if attempt > 0:
                print(f"  重试第 {attempt} 次: {warehouse1['name']} → {warehouse2['name']}")
            
            response = await agent.process_query(query)
            parsed_info = parse_distance_info(response)
            
            if parsed_info['success']:
                return {
                    'from_warehouse': warehouse1['name'],
                    'to_warehouse': warehouse2['name'],
                    'from_id': warehouse1['id'],
                    'to_id': warehouse2['id'],
                    'distance': parsed_info['distance'],
                    'duration': parsed_info['duration'],
                    'distance_km': parsed_info['distance_km'],
                    'duration_min': parsed_info['duration_min'],
                    'success': True,
                    'attempts': attempt + 1
                }
            else:
                print(f"  解析失败: {parsed_info.get('raw_response', '无响应')}")
                
        except Exception as e:
            print(f"  计算出错: {str(e)}")
            
        if attempt < max_retries - 1:
            await asyncio.sleep(1)  # 重试前等待1秒
    
    # 所有重试都失败了
    print(f"  计算失败: {warehouse1['name']} → {warehouse2['name']}")
    return {
        'from_warehouse': warehouse1['name'],
        'to_warehouse': warehouse2['name'],
        'from_id': warehouse1['id'],
        'to_id': warehouse2['id'],
        'distance': '计算失败',
        'duration': '计算失败',
        'distance_km': None,
        'duration_min': None,
        'success': False,
        'attempts': max_retries
    }

async def calculate_all_warehouse_distances(agent, warehouses):
    """计算所有仓库之间的距离（避免重复计算）"""
    print(f"开始计算 {len(warehouses)} 个仓库之间的距离...")
    
    distances = []
    # 只计算上三角矩阵，避免重复计算（A到B和B到A是相同的）
    total_pairs = len(warehouses) * (len(warehouses) - 1) // 2
    current_pair = 0
    
    for i, warehouse1 in enumerate(warehouses):
        for j, warehouse2 in enumerate(warehouses):
            if i < j:  # 只计算上三角矩阵，避免重复
                current_pair += 1
                print(f"\n进度: {current_pair}/{total_pairs} - 计算 {warehouse1['name']} ↔ {warehouse2['name']}")
                
                result = await calculate_warehouse_to_warehouse_distance(agent, warehouse1, warehouse2)
                if result:
                    distances.append(result)
                    if result['success']:
                        print(f"  ✓ 成功: {result['distance']}, {result['duration']}")
                    else:
                        print(f"  ✗ 失败: {result['distance']}")
    
    return distances

def convert_json_to_xlsx(json_file_path, xlsx_file_path, calculate_distances=True):
    """
    将JSON格式的仓库数据转换为Excel格式
    """
    # 读取JSON文件
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 1. 创建仓库基本信息工作表
    ws_basic = wb.create_sheet("仓库基本信息")
    basic_data = []
    
    for warehouse in data['warehouses']:
        basic_info = {
            '仓库ID': warehouse['id'],
            '仓库名称': warehouse['name'],
            '地址': warehouse['location']['address'],
            '经度': warehouse['location']['longitude'],
            '纬度': warehouse['location']['latitude'],
            '城市': warehouse['location']['city'],
            '行政区': warehouse['location']['district'],
            '总面积(平方米)': warehouse['capacity']['total_area'],
            '可用面积(平方米)': warehouse['capacity']['available_area'],
            '最大载重(吨)': warehouse['capacity']['max_weight'],
            '负责人': warehouse['contact']['manager'],
            '联系电话': warehouse['contact']['phone'],
            '应急电话': warehouse['contact']['emergency_phone']
        }
        basic_data.append(basic_info)
    
    df_basic = pd.DataFrame(basic_data)
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df_basic, index=False, header=True):
        ws_basic.append(r)
    
    # 设置标题样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_basic[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 自动调整列宽
    for column in ws_basic.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_basic.column_dimensions[column_letter].width = adjusted_width
    
    # 2. 创建物资详细信息工作表
    ws_resources = wb.create_sheet("物资详细信息")
    resources_data = []
    
    for warehouse in data['warehouses']:
        warehouse_id = warehouse['id']
        warehouse_name = warehouse['name']
        
        for category, items in warehouse['resources'].items():
            category_name = {
                'fire_extinguishing': '灭火设备',
                'rescue_equipment': '救援装备',
                'medical_supplies': '医疗用品',
                'communication': '通信设备',
                'evacuation': '疏散设备',
                'heavy_equipment': '重型装备',
                'logistics': '后勤保障',
                'command_center': '指挥中心'
            }.get(category, category)
            
            for item_key, item_info in items.items():
                resource_info = {
                    '仓库ID': warehouse_id,
                    '仓库名称': warehouse_name,
                    '物资类别': category_name,
                    '物资名称': item_info['type'],
                    '数量': item_info['quantity'],
                    '单位': item_info['unit'],
                    '规格说明': item_info['specification']
                }
                resources_data.append(resource_info)
    
    df_resources = pd.DataFrame(resources_data)
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df_resources, index=False, header=True):
        ws_resources.append(r)
    
    # 设置标题样式
    for cell in ws_resources[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 自动调整列宽
    for column in ws_resources.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_resources.column_dimensions[column_letter].width = adjusted_width
    
    # 3. 创建物资统计汇总工作表
    ws_summary = wb.create_sheet("物资统计汇总")
    
    # 按物资类别统计总数量
    summary_data = {}
    for warehouse in data['warehouses']:
        for category, items in warehouse['resources'].items():
            category_name = {
                'fire_extinguishing': '灭火设备',
                'rescue_equipment': '救援装备',
                'medical_supplies': '医疗用品',
                'communication': '通信设备',
                'evacuation': '疏散设备',
                'heavy_equipment': '重型装备',
                'logistics': '后勤保障',
                'command_center': '指挥中心'
            }.get(category, category)
            
            for item_key, item_info in items.items():
                item_name = item_info['type']
                quantity = item_info['quantity']
                unit = item_info['unit']
                
                if item_name not in summary_data:
                    summary_data[item_name] = {
                        '物资名称': item_name,
                        '物资类别': category_name,
                        '总数量': 0,
                        '单位': unit
                    }
                summary_data[item_name]['总数量'] += quantity
    
    df_summary = pd.DataFrame(list(summary_data.values()))
    df_summary = df_summary.sort_values(['物资类别', '物资名称'])
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)
    
    # 设置标题样式
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 自动调整列宽
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # 4. 创建元数据信息工作表
    ws_metadata = wb.create_sheet("元数据信息")
    metadata = data['metadata']
    
    metadata_info = [
        ['数据库描述', metadata['description']],
        ['版本', metadata['version']],
        ['最后更新时间', metadata['last_updated']],
        ['坐标系统', metadata['coordinate_system']],
        ['距离单位', metadata['units']['distance']],
        ['重量单位', metadata['units']['weight']],
        ['面积单位', metadata['units']['area']]
    ]
    
    for row in metadata_info:
        ws_metadata.append(row)
    
    # 设置元数据工作表样式
    for row in ws_metadata.iter_rows():
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    ws_metadata.column_dimensions['A'].width = 20
    ws_metadata.column_dimensions['B'].width = 40
    
    # 5. 计算仓库间距离（如果启用）
    distances_data = []
    if calculate_distances and len(data['warehouses']) > 1:
        print("\n开始计算仓库间距离...")
        try:
            # 异步函数来处理距离计算
            async def calculate_distances_async():
                # 创建地图代理
                agent = await create_location_agent()
                try:
                    # 计算所有仓库间距离
                    return await calculate_all_warehouse_distances(agent, data['warehouses'])
                finally:
                    # 确保断开连接
                    if hasattr(agent, 'disconnect'):
                        await agent.disconnect()
            
            # 运行异步计算
            distances_data = asyncio.run(calculate_distances_async())
            
            if distances_data:
                # 创建仓库距离工作表
                ws_distances = wb.create_sheet("仓库间距离")
                
                # 准备距离数据
                distance_rows = []
                for dist in distances_data:
                    distance_rows.append({
                        '起始仓库ID': dist['from_id'],
                        '起始仓库名称': dist['from_warehouse'],
                        '目标仓库ID': dist['to_id'],
                        '目标仓库名称': dist['to_warehouse'],
                        '距离': dist['distance'],
                        '预计时间': dist['duration'],
                        '距离(公里)': dist['distance_km'],
                        '时间(分钟)': dist['duration_min'],
                        '计算状态': '成功' if dist['success'] else '失败',
                        '尝试次数': dist['attempts']
                    })
                
                df_distances = pd.DataFrame(distance_rows)
                
                # 将数据写入工作表
                for r in dataframe_to_rows(df_distances, index=False, header=True):
                    ws_distances.append(r)
                
                # 设置标题样式
                for cell in ws_distances[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center")
                
                # 自动调整列宽
                for column in ws_distances.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws_distances.column_dimensions[column_letter].width = adjusted_width
                
                print(f"\n距离计算完成！成功计算了 {len([d for d in distances_data if d['success']])} 对仓库的距离")
            else:
                print("\n距离计算失败，未能获取任何距离数据")
                
        except Exception as e:
            print(f"\n距离计算过程中发生错误: {str(e)}")
            print("将继续保存其他数据...")
    elif not calculate_distances:
        print("\n跳过距离计算（已禁用）")
    else:
        print("\n仓库数量不足，跳过距离计算")
    
    # 保存Excel文件
    wb.save(xlsx_file_path)
    print(f"\n转换完成！Excel文件已保存到: {xlsx_file_path}")
    
    # 打印统计信息
    print(f"\n转换统计:")
    print(f"- 仓库数量: {len(data['warehouses'])}")
    print(f"- 物资条目总数: {len(resources_data)}")
    print(f"- 物资种类数: {len(summary_data)}")
    print(f"- 工作表数量: {len(wb.worksheets)}")
    if distances_data:
        successful_distances = len([d for d in distances_data if d['success']])
        print(f"- 成功计算的距离对数: {successful_distances}/{len(distances_data)}")

def main():
    import argparse
    
    # 设置命令行参数
    parser = argparse.ArgumentParser(description='将JSON格式的仓库数据转换为Excel格式，并可选择计算仓库间距离')
    parser.add_argument('--no-distances', action='store_true', help='禁用距离计算功能')
    parser.add_argument('--input', '-i', help='输入JSON文件路径')
    parser.add_argument('--output', '-o', help='输出Excel文件路径')
    
    args = parser.parse_args()
    
    # 修复路径：从项目根目录读取文件
    project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    
    # 确定输入和输出文件路径
    if args.input:
        json_file = args.input
    else:
        json_file = os.path.join(project_root, "data", "resource.json")
    
    if args.output:
        xlsx_file = args.output
    else:
        xlsx_file = os.path.join(project_root, "data", "resource.xlsx")
    
    # 确定是否计算距离
    calculate_distances = not args.no_distances
    
    # 检查JSON文件是否存在
    if not os.path.exists(json_file):
        print(f"错误: 找不到文件 {json_file}")
        return
    
    try:
        print(f"输入文件: {json_file}")
        print(f"输出文件: {xlsx_file}")
        print(f"距离计算: {'启用' if calculate_distances else '禁用'}")
        print("-" * 50)
        
        # 执行转换
        convert_json_to_xlsx(json_file, xlsx_file, calculate_distances)
        print("\n转换成功完成！")
        
    except Exception as e:
        print(f"转换过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()