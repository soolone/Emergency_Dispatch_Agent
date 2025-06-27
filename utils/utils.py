import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import json

def read_warehouse_data_from_xlsx(xlsx_file_path):
    """
    从Excel文件读取仓库数据，并返回大模型友好的格式
    
    Args:
        xlsx_file_path (str): Excel文件路径
    
    Returns:
        dict: 包含仓库信息的字典，格式化为大模型友好的结构
    """
    try:
        # 读取各个工作表
        basic_info_df = pd.read_excel(xlsx_file_path, sheet_name='仓库基本信息')
        resources_df = pd.read_excel(xlsx_file_path, sheet_name='物资详细信息')
        summary_df = pd.read_excel(xlsx_file_path, sheet_name='物资统计汇总')
        
        # 尝试读取仓库间距离数据（如果存在）
        distances_df = None
        try:
            distances_df = pd.read_excel(xlsx_file_path, sheet_name='仓库间距离')
        except:
            print("未找到仓库间距离数据表，将跳过距离信息")
        
        # 构建大模型友好的数据结构
        warehouse_data = {
            "warehouses": [],
            "resource_summary": [],
            "warehouse_distances": [],
            "total_warehouses": len(basic_info_df)
        }
        
        # 处理仓库基本信息
        for _, row in basic_info_df.iterrows():
            warehouse_info = {
                "id": row['仓库ID'],
                "name": row['仓库名称'],
                "location": {
                    "address": row['地址'],
                    "longitude": row['经度'],
                    "latitude": row['纬度'],
                    "city": row['城市'],
                    "district": row['行政区']
                },
                "capacity": {
                    "total_area": row['总面积(平方米)'],
                    "available_area": row['可用面积(平方米)'],
                    "max_weight": row['最大载重(吨)']
                },
                "contact": {
                    "manager": row['负责人'],
                    "phone": row['联系电话'],
                    "emergency_phone": row['应急电话']
                },
                "resources": []
            }
            
            # 为每个仓库添加物资信息
            warehouse_resources = resources_df[resources_df['仓库ID'] == row['仓库ID']]
            for _, resource_row in warehouse_resources.iterrows():
                resource_info = {
                    "category": resource_row['物资类别'],
                    "name": resource_row['物资名称'],
                    "quantity": resource_row['数量'],
                    "unit": resource_row['单位'],
                    "specification": resource_row['规格说明']
                }
                warehouse_info["resources"].append(resource_info)
            
            warehouse_data["warehouses"].append(warehouse_info)
        
        # 处理物资统计汇总
        for _, row in summary_df.iterrows():
            summary_info = {
                "name": row['物资名称'],
                "category": row['物资类别'],
                "total_quantity": row['总数量'],
                "unit": row['单位']
            }
            warehouse_data["resource_summary"].append(summary_info)
        
        # 处理仓库间距离数据（如果存在）
        if distances_df is not None and len(distances_df) > 0:
            for _, row in distances_df.iterrows():
                distance_info = {
                    "from_warehouse_id": row['起始仓库ID'],
                    "from_warehouse_name": row['起始仓库名称'],
                    "to_warehouse_id": row['目标仓库ID'],
                    "to_warehouse_name": row['目标仓库名称'],
                    "distance": row['距离'],
                    "duration": row['预计时间'],
                    "distance_km": row.get('距离(公里)', None),
                    "duration_min": row.get('时间(分钟)', None),
                    "status": row.get('计算状态', '未知'),
                    "attempts": row.get('尝试次数', None)
                }
                warehouse_data["warehouse_distances"].append(distance_info)
        
        return warehouse_data
        
    except Exception as e:
        print(f"读取Excel文件时发生错误: {e}")
        return None

def format_warehouse_data_for_llm(warehouse_data):
    """
    将仓库数据格式化为更适合大模型理解的文本格式
    
    Args:
        warehouse_data (dict): 仓库数据字典
    
    Returns:
        tuple: (仓库信息文本, 仓库间距离文本)
    """
    if not warehouse_data:
        return "无法获取仓库数据", ""
    
    # 仓库信息文本
    warehouse_text = f"成都市应急物资仓库信息概览\n\n"
    warehouse_text += f"总仓库数量: {warehouse_data['total_warehouses']}个\n\n"
    
    # 仓库详细信息
    warehouse_text += "=== 仓库详细信息 ===\n"
    for warehouse in warehouse_data['warehouses']:
        warehouse_text += f"\n【{warehouse['name']}】\n"
        warehouse_text += f"- 仓库ID: {warehouse['id']}\n"
        warehouse_text += f"- 地址: {warehouse['location']['address']}\n"
        warehouse_text += f"- 位置: {warehouse['location']['city']}{warehouse['location']['district']}\n"
        warehouse_text += f"- 坐标: ({warehouse['location']['longitude']}, {warehouse['location']['latitude']})\n"
        warehouse_text += f"- 容量: 总面积{warehouse['capacity']['total_area']}㎡, 可用面积{warehouse['capacity']['available_area']}㎡, 最大载重{warehouse['capacity']['max_weight']}吨\n"
        warehouse_text += f"- 联系人: {warehouse['contact']['manager']} (电话: {warehouse['contact']['phone']}, 应急: {warehouse['contact']['emergency_phone']})\n"
        
        if warehouse['resources']:
            warehouse_text += f"- 主要物资:\n"
            # 按类别分组显示物资
            resources_by_category = {}
            for resource in warehouse['resources']:
                category = resource['category']
                if category not in resources_by_category:
                    resources_by_category[category] = []
                resources_by_category[category].append(resource)
            
            for category, resources in resources_by_category.items():
                warehouse_text += f"  * {category}:\n"
                for resource in resources:
                    warehouse_text += f"    - {resource['name']}: {resource['quantity']}{resource['unit']} ({resource['specification']})\n"
    
    # 物资统计汇总
    warehouse_text += "\n=== 物资统计汇总 ===\n"
    resources_by_category = {}
    for resource in warehouse_data['resource_summary']:
        category = resource['category']
        if category not in resources_by_category:
            resources_by_category[category] = []
        resources_by_category[category].append(resource)
    
    for category, resources in resources_by_category.items():
        warehouse_text += f"\n【{category}】\n"
        for resource in resources:
            warehouse_text += f"- {resource['name']}: {resource['total_quantity']}{resource['unit']}\n"
    
    # 仓库间距离文本
    distance_text = ""
    if 'warehouse_distances' in warehouse_data and warehouse_data['warehouse_distances']:
        distance_pairs = []
        for distance in warehouse_data['warehouse_distances']:
            if distance['status'] == '成功':
                distance_info = f"{distance['from_warehouse_name']}-{distance['to_warehouse_name']}：{distance['distance_km']}公里"
                if distance['duration']:
                    distance_info += f", 预计时间: {distance['duration']}"
                distance_pairs.append(distance_info)
        
        if distance_pairs:
            distance_text = "\n".join(distance_pairs)
    
    return warehouse_text, distance_text

def convert_json_to_xlsx(json_file_path, xlsx_file_path):
    """
    将JSON格式的仓库数据转换为Excel格式
    """
    # 读取JSON文件
    with open(json_file_path, 'r', encoding='utf-8') as f:
        data = json.load(f)
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 删除默认工作表
    wb.remove(wb.active)
    
    # 1. 创建仓库基本信息工作表
    ws_basic = wb.create_sheet("仓库基本信息")
    basic_data = []
    
    for warehouse in data['warehouses']:
        basic_info = {
            '仓库ID': warehouse['id'],
            '仓库名称': warehouse['name'],
            '地址': warehouse['location']['address'],
            '经度': warehouse['location']['longitude'],
            '纬度': warehouse['location']['latitude'],
            '城市': warehouse['location']['city'],
            '行政区': warehouse['location']['district'],
            '总面积(平方米)': warehouse['capacity']['total_area'],
            '可用面积(平方米)': warehouse['capacity']['available_area'],
            '最大载重(吨)': warehouse['capacity']['max_weight'],
            '负责人': warehouse['contact']['manager'],
            '联系电话': warehouse['contact']['phone'],
            '应急电话': warehouse['contact']['emergency_phone']
        }
        basic_data.append(basic_info)
    
    df_basic = pd.DataFrame(basic_data)
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df_basic, index=False, header=True):
        ws_basic.append(r)
    
    # 设置标题样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws_basic[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 自动调整列宽
    for column in ws_basic.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_basic.column_dimensions[column_letter].width = adjusted_width
    
    # 2. 创建物资详细信息工作表
    ws_resources = wb.create_sheet("物资详细信息")
    resources_data = []
    
    for warehouse in data['warehouses']:
        warehouse_id = warehouse['id']
        warehouse_name = warehouse['name']
        
        for category, items in warehouse['resources'].items():
            category_name = {
                'fire_extinguishing': '灭火设备',
                'rescue_equipment': '救援装备',
                'medical_supplies': '医疗用品',
                'communication': '通信设备',
                'evacuation': '疏散设备',
                'heavy_equipment': '重型装备',
                'logistics': '后勤保障',
                'command_center': '指挥中心'
            }.get(category, category)
            
            for item_key, item_info in items.items():
                resource_info = {
                    '仓库ID': warehouse_id,
                    '仓库名称': warehouse_name,
                    '物资类别': category_name,
                    '物资名称': item_info['type'],
                    '数量': item_info['quantity'],
                    '单位': item_info['unit'],
                    '规格说明': item_info['specification']
                }
                resources_data.append(resource_info)
    
    df_resources = pd.DataFrame(resources_data)
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df_resources, index=False, header=True):
        ws_resources.append(r)
    
    # 设置标题样式
    for cell in ws_resources[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 自动调整列宽
    for column in ws_resources.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws_resources.column_dimensions[column_letter].width = adjusted_width
    
    # 3. 创建物资统计汇总工作表
    ws_summary = wb.create_sheet("物资统计汇总")
    
    # 按物资类别统计总数量
    summary_data = {}
    for warehouse in data['warehouses']:
        for category, items in warehouse['resources'].items():
            category_name = {
                'fire_extinguishing': '灭火设备',
                'rescue_equipment': '救援装备',
                'medical_supplies': '医疗用品',
                'communication': '通信设备',
                'evacuation': '疏散设备',
                'heavy_equipment': '重型装备',
                'logistics': '后勤保障',
                'command_center': '指挥中心'
            }.get(category, category)
            
            for item_key, item_info in items.items():
                item_name = item_info['type']
                quantity = item_info['quantity']
                unit = item_info['unit']
                
                if item_name not in summary_data:
                    summary_data[item_name] = {
                        '物资名称': item_name,
                        '物资类别': category_name,
                        '总数量': 0,
                        '单位': unit
                    }
                summary_data[item_name]['总数量'] += quantity
    
    df_summary = pd.DataFrame(list(summary_data.values()))
    df_summary = df_summary.sort_values(['物资类别', '物资名称'])
    
    # 将数据写入工作表
    for r in dataframe_to_rows(df_summary, index=False, header=True):
        ws_summary.append(r)
    
    # 设置标题样式
    for cell in ws_summary[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # 自动调整列宽
    for column in ws_summary.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # 4. 创建元数据信息工作表
    ws_metadata = wb.create_sheet("元数据信息")
    metadata = data['metadata']
    
    metadata_info = [
        ['数据库描述', metadata['description']],
        ['版本', metadata['version']],
        ['最后更新时间', metadata['last_updated']],
        ['坐标系统', metadata['coordinate_system']],
        ['距离单位', metadata['units']['distance']],
        ['重量单位', metadata['units']['weight']],
        ['面积单位', metadata['units']['area']]
    ]
    
    for row in metadata_info:
        ws_metadata.append(row)
    
    # 设置元数据工作表样式
    for row in ws_metadata.iter_rows():
        row[0].font = Font(bold=True)
        row[0].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
    
    ws_metadata.column_dimensions['A'].width = 20
    ws_metadata.column_dimensions['B'].width = 40
    
    # 保存Excel文件
    wb.save(xlsx_file_path)
    print(f"转换完成！Excel文件已保存到: {xlsx_file_path}")
    
    # 打印统计信息
    print(f"\n转换统计:")
    print(f"- 仓库数量: {len(data['warehouses'])}")
    print(f"- 物资条目总数: {len(resources_data)}")
    print(f"- 物资种类数: {len(summary_data)}")
    print(f"- 工作表数量: {len(wb.worksheets)}")

if __name__=="__main__":
    warehouse_data = read_warehouse_data_from_xlsx("e:/proj_lab/Emergency_Agent/data/resource.xlsx")
    warehouse_text, distance_text = format_warehouse_data_for_llm(warehouse_data)
    print("=== 仓库信息 ===")
    print(warehouse_text)
    print("\n=== 仓库间距离 ===")
    print(distance_text)